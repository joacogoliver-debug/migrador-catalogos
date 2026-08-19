"""
Armado del entregable: un ZIP organizado por producto, listo para entregar a la
distribuidora nueva.

Estructura:

    Artista - Migracion 2026-08-13/
    ├── _LEEME.txt                     ← qué es esto y cómo está organizado
    ├── _Catalogo completo.xlsx        ← planilla maestra de todos los productos
    ├── _Reporte de migracion.txt      ← qué salió, qué faltó y por qué
    ├── 2019 - Nombre del Album [UPC]/
    │   ├── portada.jpg
    │   ├── datos.xlsx                 ← sólo este producto, con ISRCs
    │   ├── 01 - Primer Tema.flac
    │   └── 02 - Segundo Tema.flac
    └── 2021 - Nombre del Single [UPC]/
        └── ...

Por qué por producto y no por tipo de archivo: en una migración cada producto se
entrega como una unidad (un UPC, una portada, sus audios). Así cada carpeta ya
queda lista para subir, sin tener que cruzar tres carpetas distintas para armar
un release. El UPC en el nombre evita confundir un álbum con su reedición.

El ZIP se escribe a disco y no en memoria: un catálogo mediano en FLAC son
varios GB y no entra en RAM.
"""

import os
import re
import unicodedata
import zipfile
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from audio import ETIQUETA_LOSSLESS, FORMATOS_LOSSLESS

NAVY = "1F3864"
GRIS = "F2F2F2"
AMBAR = "FFF2CC"


def _slug_archivo(s, maxlen=80):
    """Nombre de archivo seguro en Windows/macOS/Linux."""
    s = unicodedata.normalize("NFD", s or "").encode("ascii", "ignore").decode("ascii")
    s = re.sub(r'[<>:"/\\|?*]', "", s)
    s = re.sub(r"[\x00-\x1f]", "", s)
    s = re.sub(r"\s+", " ", s).strip(" .")
    return (s[:maxlen].strip(" .") or "sin-titulo")


def _fuente_corta(t):
    """Etiqueta corta de fuente/calidad para la planilla."""
    if not t.get("audio_path"):
        return "sin audio"
    fmt = (t.get("audio_format") or "").lstrip(".")
    return f"{'LOSSLESS' if t.get('audio_format') in FORMATOS_LOSSLESS else 'LOSSY'} ({fmt})"


# ============================================================
# Planillas
# ============================================================

COLUMNAS = [
    ("Producto", 34), ("Tipo", 8), ("Año", 6), ("UPC", 15),
    ("#", 4), ("Track", 34), ("ISRC", 14),
    ("Duración", 9), ("Sello", 22), ("Distribuidora", 22),
    ("Fuente / Calidad", 26), ("Archivo", 30), ("Reproducciones", 14),
    ("URL YouTube", 30),
]


def _encabezado(ws, titulo, subtitulo=""):
    ws["A1"] = titulo
    ws["A1"].font = Font(size=14, bold=True, color=NAVY)
    if subtitulo:
        ws["A2"] = subtitulo
        ws["A2"].font = Font(size=9, color="666666")
    fila = 4
    for i, (nombre, ancho) in enumerate(COLUMNAS, 1):
        c = ws.cell(row=fila, column=i, value=nombre)
        c.font = Font(size=10, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = ancho
    ws.freeze_panes = f"A{fila + 1}"
    return fila + 1


def _mmss(seg):
    seg = int(seg or 0)
    return f"{seg // 60}:{seg % 60:02d}"


def _filas_producto(ws, fila, p, con_archivo=True):
    for t in p["tracks"]:
        fuente = _fuente_corta(t)
        valores = [
            p["title"], p["kind"], p.get("release_year", ""), p.get("upc", ""),
            t.get("track_number", ""), t.get("track", ""), t.get("isrc", ""),
            _mmss(t.get("duration_s")), p.get("label", ""), p.get("distributor", ""),
            fuente,
            os.path.basename(t["audio_path"]) if (con_archivo and t.get("audio_path")) else "",
            t.get("views", 0), t.get("url", ""),
        ]
        for i, v in enumerate(valores, 1):
            c = ws.cell(row=fila, column=i, value=v)
            c.alignment = Alignment(vertical="center")
            # Resaltamos en ámbar lo que NO es apto para entrega, para que no se
            # cuele un lossy en una entrega por distracción.
            if i == 11 and fuente != "sin audio" and not fuente.startswith("LOSSLESS"):
                c.fill = PatternFill("solid", fgColor=AMBAR)
            if i == 11 and fuente == "sin audio":
                c.font = Font(size=10, color="C00000")
        fila += 1
    return fila


def planilla_maestra_bytes(productos, artista):
    """Excel con todo el catálogo seleccionado."""
    from io import BytesIO
    wb = Workbook()
    ws = wb.active
    ws.title = "Catálogo"
    hoy = date.today().isoformat()
    fila = _encabezado(
        ws, f"{artista} — Catálogo para migración",
        f"Generado el {hoy} · {len(productos)} productos · "
        f"{sum(p['track_count'] for p in productos)} tracks",
    )
    for p in productos:
        fila = _filas_producto(ws, fila, p)
    ws.auto_filter.ref = f"A4:{get_column_letter(len(COLUMNAS))}{fila - 1}"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ============================================================
# Hoja de ingesta — CSV para cargar en la distribuidora nueva
# ============================================================
#
# Las distribuidoras ingestan por planilla propia o por DDEX ERN. DDEX quedó
# afuera a propósito: emitir ERN válido requiere ser parte registrada de DDEX con
# un DPID propio, así que un XML "casi DDEX" sería peor que no darlo (se rechaza
# igual y da falsa sensación de que está listo). En su lugar damos un CSV con las
# columnas estándar que aceptan o mapean casi todas, y marcamos explícitamente lo
# que sólo puede completar el dueño del catálogo.

MARCA_COMPLETAR = "<<COMPLETAR>>"

COLUMNAS_INGESTA = [
    # --- nivel release ---
    "UPC", "Release Title", "Release Artist", "Release Type", "Release Date",
    "Label", "P Line", "C Line", "Genre", "Language", "Territories",
    # --- nivel track ---
    "Disc Number", "Track Number", "ISRC", "Track Title", "Track Artist",
    "Duration", "Explicit", "Composer", "Publisher", "Lyrics Language",
    # --- referencia interna ---
    "Audio File", "Cover File", "Source Quality", "YouTube URL",
]


def hoja_ingesta_csv(productos, artista):
    """CSV con las columnas estándar de ingesta, una fila por track.

    Lo que sabemos va completo; lo que no puede salir de YouTube ni de las APIs
    públicas (género, explicit, compositores, editoriales) queda marcado con
    <<COMPLETAR>> en vez de vacío o inventado, así se ve de una qué falta.
    """
    import csv
    from io import StringIO

    buf = StringIO()
    w = csv.writer(buf, lineterminator="\n")
    w.writerow(COLUMNAS_INGESTA)

    for p in productos:
        anio = p.get("release_year") or ""
        sello = p.get("label") or MARCA_COMPLETAR
        # La línea ℗ se arma con lo que trae YouTube; si falta el año no la
        # inventamos.
        p_line = f"{anio} {sello}".strip() if anio and p.get("label") else MARCA_COMPLETAR
        for t in p["tracks"]:
            w.writerow([
                p.get("upc") or MARCA_COMPLETAR,
                p.get("title", ""),
                artista,
                p.get("kind", ""),
                p.get("release_date") or (f"{anio}-01-01" if anio else MARCA_COMPLETAR),
                sello,
                p_line,
                MARCA_COMPLETAR,            # C Line: no sale de YouTube
                MARCA_COMPLETAR,            # Genre
                MARCA_COMPLETAR,            # Language
                "Worldwide",
                t.get("tidal", {}).get("volume_number") if t.get("tidal") else 1,
                t.get("track_number") or "",
                t.get("isrc") or MARCA_COMPLETAR,
                t.get("track", ""),
                artista,
                _mmss(t.get("duration_s")),
                MARCA_COMPLETAR,            # Explicit
                MARCA_COMPLETAR,            # Composer
                MARCA_COMPLETAR,            # Publisher
                MARCA_COMPLETAR,            # Lyrics Language
                os.path.basename(t["audio_path"]) if t.get("audio_path") else "",
                "portada.jpg" if p.get("cover_bytes") else "",
                _fuente_corta(t),
                t.get("url", ""),
            ])
    return buf.getvalue()


def planilla_producto_bytes(p, artista):
    """Excel de un solo producto, para que viaje dentro de su carpeta."""
    from io import BytesIO
    wb = Workbook()
    ws = wb.active
    ws.title = "Producto"
    fila = _encabezado(
        ws, f"{artista} — {p['title']}",
        f"{p['kind'].upper()} · {p.get('release_year', 's/f')} · "
        f"UPC {p.get('upc') or '(sin UPC)'} · {p['track_count']} tracks",
    )
    _filas_producto(ws, fila, p)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ============================================================
# Reporte y leeme
# ============================================================

def reporte_texto(productos, artista, entorno=None, con_tidal=False):
    """Reporte honesto de qué se pudo migrar y qué no. Es la pieza que evita
    sorpresas: dice producto por producto qué falta y por qué."""
    L = []
    tracks = [t for p in productos for t in p["tracks"]]
    aptos = [t for t in tracks if (t.get("audio_format") or "") in FORMATOS_LOSSLESS]
    ref = [t for t in tracks if t.get("audio_path") and t not in aptos]
    sin_audio = [t for t in tracks if not t.get("audio_path")]

    L.append(f"REPORTE DE MIGRACIÓN — {artista}")
    L.append(f"Generado: {date.today().isoformat()}")
    L.append("=" * 68)
    L.append("")
    L.append(f"Productos seleccionados : {len(productos)}")
    L.append(f"Tracks totales          : {len(tracks)}")
    L.append(f"  Con ISRC              : {sum(1 for t in tracks if t.get('isrc'))}")
    L.append(f"Productos con UPC       : {sum(1 for p in productos if p.get('upc'))}")
    L.append(f"Portadas obtenidas      : {sum(1 for p in productos if p.get('cover_bytes'))}/{len(productos)}")
    L.append("")
    L.append("AUDIO")
    L.append(f"  Aptos para entrega (FLAC lossless) : {len(aptos)}")
    L.append(f"  Sólo referencia (lossy)            : {len(ref)}")
    L.append(f"  Sin audio                          : {len(sin_audio)}")
    L.append("")

    if not con_tidal:
        L.append("! No se conectó una cuenta de Tidal, así que NO hay audio apto para")
        L.append("  entrega. Todo el audio de este paquete es referencia lossy de")
        L.append("  YouTube. Para una entrega real hace falta el máster original.")
        L.append("")
    elif ref:
        L.append("! Algunos tracks bajaron en AAC y no en FLAC: Tidal no tiene máster")
        L.append("  lossless para esas grabaciones. Están marcados como lossy y NO")
        L.append("  son aptos para entrega — hay que pedir el máster al sello/artista.")
        L.append("")

    L.append("PENDIENTES POR PRODUCTO")
    L.append("-" * 68)
    hay_pendientes = False
    for p in productos:
        faltas = []
        if not p.get("upc"):
            faltas.append("sin UPC")
        if not p.get("cover_bytes"):
            faltas.append(f"sin portada ({p.get('cover_status', 'no buscada')})")
        sin = [t for t in p["tracks"] if not t.get("audio_path")]
        if sin:
            faltas.append(f"{len(sin)}/{p['track_count']} tracks sin audio")
            # El motivo concreto por track: sirve para saber si hay que buscar
            # otra fuente o si el video simplemente ya no está.
            for t in sin:
                if t.get("audio_error"):
                    faltas.append(f"    · {t.get('track', '')[:40]}: {t['audio_error']}")
        lossy = [t for t in p["tracks"]
                 if t.get("audio_path") and (t.get("audio_format") or "") not in FORMATOS_LOSSLESS]
        if lossy:
            faltas.append(f"{len(lossy)} tracks sólo en calidad de referencia")
        sin_isrc = [t for t in p["tracks"] if not t.get("isrc")]
        if sin_isrc:
            faltas.append(f"{len(sin_isrc)} tracks sin ISRC")
        if p.get("order_unconfirmed"):
            faltas.append("orden de tracks sin confirmar (estimado por fecha de subida)")
        if faltas:
            hay_pendientes = True
            L.append(f"  {p['folder']}")
            for f in faltas:
                L.append(f"      - {f}")
    if not hay_pendientes:
        L.append("  (ninguno: todos los productos quedaron completos)")

    if entorno:
        L.append("")
        L.append("ENTORNO")
        L.append(f"  ffmpeg: {'sí' if entorno.get('ffmpeg') else 'NO'} · "
                 f"tiddl: {'sí' if entorno.get('tiddl') else 'NO'} · "
                 f"yt-dlp: {'sí' if entorno.get('yt_dlp') else 'NO'}")
    return "\n".join(L) + "\n"


LEEME = """CÓMO ESTÁ ORGANIZADO ESTE PAQUETE
=================================

Una carpeta por producto (álbum / EP / single). Cada una trae:

  portada.jpg   La portada en la resolución más alta que tenía Apple Music
                (hasta 3000x3000).
  datos.xlsx    Los datos de ese producto: tracks, ISRC, UPC, sello, duración.
  NN - Tema.ext Los audios, numerados en el orden del release.

En la raíz:

  _Catalogo completo.xlsx      Todos los productos en una sola planilla.
  _Hoja de ingesta.csv         El archivo para cargar en la distribuidora nueva.
  _Validacion pre-entrega.txt  Qué va a ser rechazado y qué conviene revisar.
  _Reporte de migracion.txt    Qué se pudo obtener y qué quedó pendiente.

EMPEZÁ POR LA VALIDACIÓN
------------------------
Abrí primero "_Validacion pre-entrega.txt". Separa dos cosas:

  ERRORES  La distribuidora los rechaza (código con formato inválido, dígito
           verificador mal, código duplicado, portada chica o no cuadrada).
           Hay que corregirlos antes de entregar.

  AVISOS   Pasan la ingesta pero conviene revisarlos (falta un ISRC o un UPC y
           se va a asignar uno nuevo, un título arrastra texto de YouTube).

SOBRE LA HOJA DE INGESTA
------------------------
"_Hoja de ingesta.csv" trae las columnas estándar que aceptan o mapean casi
todas las distribuidoras. Lo que se pudo relevar viene completo. Lo que no puede
salir de fuentes públicas está marcado con <<COMPLETAR>>:

  Genre, Language, Explicit, Composer, Publisher, C Line

Esos campos los tiene que llenar el dueño del catálogo — están marcados en vez
de vacíos o inventados justamente para que no pasen desapercibidos.

SOBRE LA CALIDAD DEL AUDIO — LEER ANTES DE ENTREGAR
---------------------------------------------------
La columna "Fuente / Calidad" de las planillas dice, track por track, de dónde
salió el audio:

  LOSSLESS (flac)  Máster lossless de Tidal. Apto para entregar.

  LOSSY (m4a/opus/webm)  Audio ya comprimido. Sirve como referencia, inventario
                o verificación, pero NO es apto para entregar a una
                distribuidora: se subiría con pérdida de calidad. Estos casos
                están resaltados en ámbar en la planilla.

Si un track figura como LOSSY, hay que conseguir el máster original con el
artista o el sello antes de la entrega. El reporte lista exactamente cuáles.

OTROS DATOS QUE SON ESTIMADOS
-----------------------------
YouTube no declara todo lo que necesita una ficha de release, así que dos campos
son aproximaciones y conviene verificarlos:

  Tipo (single/EP/álbum)  Se deduce de la cantidad de tracks (1-3 single,
                          4-6 EP, 7+ álbum). Un EP corto puede figurar como
                          single.

  Orden de los tracks     Cuando se pudo cruzar con Tidal por ISRC, el número
                          de track es el real. Si no, es un estimado por fecha
                          de subida y el reporte lo marca como "sin confirmar".
"""


# ============================================================
# ZIP
# ============================================================

def build_zip(productos, artista, out_path, entorno=None, con_tidal=False,
              incluir_planilla=True, incluir_audio=True, incluir_portadas=True,
              log=print):
    """Arma el ZIP del entregable en `out_path`. Devuelve (ruta, bytes).

    Se escribe directo a disco porque un catálogo en FLAC son varios GB.
    """
    raiz = f"{_slug_archivo(artista)} - Migracion {date.today().isoformat()}"
    # ZIP_STORED para el audio: FLAC y Opus ya están comprimidos, deflate
    # gastaría CPU sin ganar espacio. Sí comprimimos planillas y texto.
    with zipfile.ZipFile(out_path, "w", zipfile.ZIP_DEFLATED, allowZip64=True) as z:
        # Los .txt van con BOM (utf-8-sig): los abre gente en Windows y sin BOM
        # algunos editores viejos muestran los acentos rotos.
        z.writestr(f"{raiz}/_LEEME.txt", LEEME.encode("utf-8-sig"))
        z.writestr(f"{raiz}/_Reporte de migracion.txt",
                   reporte_texto(productos, artista, entorno, con_tidal).encode("utf-8-sig"))

        # La validación va siempre: es lo que evita que la entrega se rechace.
        import validar as V
        res_val = V.validar(productos, artista)
        z.writestr(f"{raiz}/_Validacion pre-entrega.txt",
                   V.reporte_validacion(res_val, artista).encode("utf-8-sig"))

        if incluir_planilla:
            z.writestr(f"{raiz}/_Catalogo completo.xlsx",
                       planilla_maestra_bytes(productos, artista))
            # CSV de ingesta: es el archivo que se carga en la distribuidora.
            z.writestr(f"{raiz}/_Hoja de ingesta.csv",
                       hoja_ingesta_csv(productos, artista).encode("utf-8-sig"))

        for p in productos:
            carpeta = f"{raiz}/{p['folder']}"
            if incluir_planilla:
                z.writestr(f"{carpeta}/datos.xlsx", planilla_producto_bytes(p, artista))
            if incluir_portadas and p.get("cover_bytes"):
                z.writestr(f"{carpeta}/portada.jpg", p["cover_bytes"])

            if incluir_audio:
                for t in p["tracks"]:
                    ruta = t.get("audio_path")
                    if not ruta or not os.path.exists(ruta):
                        continue
                    n = t.get("track_number") or 0
                    ext = t.get("audio_format") or os.path.splitext(ruta)[1]
                    nombre = f"{n:02d} - {_slug_archivo(t.get('track'))}{ext}"
                    # Los lossy van marcados en el nombre del archivo: es la
                    # última barrera para que no se entreguen por error.
                    if t.get("audio_format") not in FORMATOS_LOSSLESS:
                        nombre = f"{n:02d} - {_slug_archivo(t.get('track'))} [REFERENCIA-LOSSY]{ext}"
                    z.write(ruta, f"{carpeta}/{nombre}", compress_type=zipfile.ZIP_STORED)
            log(f"[zip] {p['folder']}")

    tam = os.path.getsize(out_path)
    log(f"[zip] listo: {out_path} ({tam / 1e6:.1f} MB)")
    return out_path, tam
